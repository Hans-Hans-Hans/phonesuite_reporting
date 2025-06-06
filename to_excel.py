from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def create_workbook(data: list):
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Devices"

    # Define headers
    headers = [
        "Device ID", "Type", "Name", "Status", "IP", "Port", "Description",
        "Assignments", "Pickup Group", "SLA", "Is Trunk?"
    ]

    # Write headers
    ws.append(headers)

    # Write each row of device data
    for row in data:
        ws.append(row)

    # Optional: Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column number
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Save the workbook
    output_file = "status_report.xlsx"
    wb.save(output_file)
    print(f"Saved Excel file to {output_file}")